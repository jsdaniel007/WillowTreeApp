#!/usr/bin/env python
import openpyxl
import os

os.chdir('outputDocs')
wb = openpyxl.Workbook()

# Get the 'first' worksheet
ws = wb.active
ws.title = "Chris' Title"

ws1 = wb.create_sheet("Chris' Sheet") # append new sheet, second param can denote pos

#ws.sheet_properties.tabColor = "1072BA"

ws3 = wb["Chris' Title"] # Grabbing a worksheet by name

#iterate through our 2 worksheets
for sheet in wb:
	print(sheet.title)

# Data Modification
cell_example = ws['A1']
cell_range = ws['A1':'C2']
columnC = ws['C']
row1 = ws[10]
row_range = ws[5:10]

# Storage and Saving
cell = ws['A5']
cell2 = ws['A6']
cell.value = 'hello, example!'
cell2.value = 3.14

# Save to a file
wb.save('example.xlsx')
