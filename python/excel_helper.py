#!/usr/bin/env python
import openpyxl
import enum
import os

"""
Purpose: to generalize some excel functions that inventory/product python files
will use such as:
- Creating the initial column titles

"""

# This can be hashed at the start of the class
class Sheet(enum.Enum):
	NONE = "None"
	TR = "Transactions"
	PR = "Products"

# Base Class for sheet interactions
class ExcelHelper():
	# Columns for the screen
	m_WorkbookFile = "" # With the extension
	m_TransactionColumns = ['Date', 'Quantity', 'Product_Name', 'Customer/UserName', 'Net_Price']
	m_ProductColumns = ['Product_Name', 'Total_Quantity', 'Retail_Price', 'Shipping_Cost']


	# Create a workbook for with initial settings
	def __init__(self, workbookName):

		# Create Workbook and create initial sheets
		self.workbookObj = openpyxl.Workbook() # Create a wb object
		self.m_WorkbookFile = workbookName + '.xlsx'

		# Sheet Creation
		self.workbookObj.active.title = Sheet.TR.value # first page rename
		self.workbookObj.create_sheet(Sheet.PR.value)

		# Define Sheet Columns
		self.defineColumns(Sheet.TR.value, self.m_TransactionColumns)
		self.defineColumns(Sheet.PR.value, self.m_ProductColumns)

		# Save the workbook to the previously declared location
		self.save()

	# Travel across the top row and label 'screenName' sheet cells, w/ columns list elements
	def defineColumns(self, sheetName, columnsList):
		sheet = self.workbookObj[sheetName]

		# Load up columns with titles
		i = 0
		for col in sheet.iter_cols(min_row=1, max_col=len(columnsList)):
			for cell in col:
				cell.value = columnsList[i]
				i += 1

		self.m_RowCount += 1

	# Add a row of {dataList} contents to {sheetName} cells, overwriting previous data
	def addRow(self, sheetName, dataList):
		sheet = self.workbookObj[sheetName]

		# Load up columns with titles
		i = 0
		for col in sheet.iter_cols(min_row=0, max_col=len(dataList)):
			for cell in col:
				cell.value = dataList[i]
				i += 1
		#add a row to the count
		


	# TODO: Testing needed to confirm it works, depending on row placement
	# Add a row of {dataList} contents to {sheetName} cells, leaving matching data alone
	def editRow(self, sheetName, dataList):
		sheet = self.workbookObj[sheetName]

		# Load up columns with titles, unless it's the same as the sent in dataList
		i = 0
		# TODO: Fix row placement?
		for col in sheet.iter_cols(min_row=1, max_col=len(dataList)):
			for cell in col:
				if not cell.value == dataList[i]:
					cell.value = sheetName[i]
				i += 1

	def save(self):
		self.workbookObj.save(self.m_WorkbookFile)

if __name__ == '__main__':
	workbookName = 'sheets/example'
	# This is assuming the data is already validated
	transData = ['05/20/2020', 99, "Blue Bottle", "supercool100", 10.99]
	productData = ["Blue Bottle", 250, 10.99, 2.99]


	helperObj = ExcelHelper(workbookName)

	helperObj.addRow(Sheet.TR.value, transData)
	helperObj.addRow(Sheet.PR.value, productData)
	helperObj.save()
